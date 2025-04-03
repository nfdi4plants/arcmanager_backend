import pandas as pd
from json import loads
import numpy as np
import os
from fastapi import HTTPException
import openpyxl

from app.models.gitlab.input import sheetContent

from fsspreadsheet.xlsx import Xlsx
from fsspreadsheet.workbook import FsWorkbook, FsWorksheet


def getRowIndex(name: str, worksheet: FsWorksheet):
    firstColumn = FsWorksheet.get_column_at(1, worksheet)

    for x in range(1, firstColumn.MaxRowIndex):
        if name == firstColumn.Item(x).Value:
            return x

    return -1


# sanitize input
def sanitizeInput(input: str | list) -> str | list:
    if type(input) is list:
        return [sanitizeInput(entry) for entry in input]

    if type(input) is str:
        return input.replace("<", "&lt;").replace(">", "&gt;")
    return input


# reads out the given file and sends the content as json back
def readIsaFile(path: str, type: str):
    # initiate isaFile structure
    isaFile: pd.DataFrame

    # match the correct sheet name with the given type of isa
    match type:
        case "investigation":
            sheetName = sheetName2 = "isa_investigation"

        case "study":
            sheetName = "Study"

            # the intended name stated in the arc specification
            sheetName2 = "isa_study"

        case "assay":
            sheetName = "Assay"

            # the intended name stated in the arc specification
            sheetName2 = "isa_assay"

        case "datamap":
            return getSwateSheets(path, "datamap")

        case other:
            sheetName = sheetName2 = ""

    # read the file
    try:
        isaFile = pd.read_excel(path, sheet_name=sheetName, engine="openpyxl")
    except:
        try:
            isaFile = pd.read_excel(path, sheet_name=sheetName2, engine="openpyxl")
        except:
            # if none matches, just read the file with default values
            isaFile = pd.read_excel(path, 0, engine="openpyxl")

    # parse the dataframe into json and return it
    return loads(isaFile.to_json(orient="split"))


# replaces the old content of the file with the new content
def writeIsaFile(path: str, type: str, newContent, repoId: int, location: str):
    # construct the path with the given values (e.g. .../freiburg-33/isa.investigation.xlsx)
    pathName = f"{os.environ.get('BACKEND_SAVE')}{location}-{repoId}/{path}"

    # match the correct sheet name with the given type of isa

    sheetIndex = -1
    match type:
        case "investigation":
            sheetName = "isa_investigation"

        case "study":
            sheetName2 = "Study"

            # the intended name stated in the arc specification
            sheetName = "isa_study"

        case "assay":
            sheetName2 = "Assay"

            # the intended name stated in the arc specification
            sheetName = "isa_assay"

        case "datamap":
            sheetName = "isa_datamap"

            sheetName2 = "Datamap"

        case other:
            sheetName = sheetName2 = ""

    # read the file
    try:
        isaFile = pd.read_excel(pathName, sheet_name=sheetName, engine="openpyxl")

    except:
        try:
            sheetName = sheetName2
            isaFile = pd.read_excel(pathName, sheet_name=sheetName2, engine="openpyxl")
        except:
            sheetName = 0
            isaFile = pd.read_excel(pathName, 0, engine="openpyxl")

    try:
        importIsa = Xlsx.from_xlsx_file(pathName)
        # get the index of the correct sheet
        for i, sheet in enumerate(FsWorkbook.get_worksheets(importIsa)):
            if sheet.name == sheetName or sheetName == sheetName2:
                sheetIndex = i
                break

        if sheetIndex > -1:
            sheetData = FsWorkbook.get_worksheets(importIsa)[sheetIndex]
        else:
            sheetData = FsWorkbook.get_worksheets(importIsa)[0]

        rowIndex = getRowIndex(newContent[0], sheetData)

        for x in range(1, len(newContent)):
            if newContent[x] != None and newContent[x] != "":
                sheetData.SetValueAt(sanitizeInput(newContent[x]), rowIndex, x + 1)
        try:
            importIsa.RemoveWorksheet(sheetName)
        except:
            importIsa.RemoveWorksheet(sheetName2)

        importIsa.AddWorksheet(sheetData)

        Xlsx.to_file(
            pathName,
            importIsa,
        )

        return FsWorksheet.get_cell_at(rowIndex, 2, sheetData).Value
    except:

        # replace nan values with empty strings
        isaFile = isaFile.fillna("")
        if newContent[0] != "":
            # get the id of the row to edit
            id = isaFile.index[
                isaFile[isaFile[0:1].columns[0]] == newContent[0]
            ].values[0]

            # get the current content to know what to replace
            oldContent = isaFile[id : id + 1]

            # Here we replace every entry in the corresponding field with the new value (column by column)
            for x in range(1, len(newContent)):
                # if there are new fields in newContent insert a new column "Unnamed: number" with empty fields
                if x > oldContent.count(axis="columns").values[0] - 1:
                    try:
                        isaFile.insert(x, "Unnamed: " + str(x), "")
                        # add the new field to old content to extent its length
                        oldContent.insert(x, "Unnamed: " + str(x), "")
                    except:
                        isaFile.insert(x, "Unnamed")
                        oldContent.insert(x, "Unnamed")

                # get the name of the current column
                columnName = isaFile[id : id + 1].columns[x]

                # read out the value on the row with the given id and the current column and replace it with the new value
                isaFile[id : id + 1].at[id, columnName] = (
                    isaFile[id : id + 1]
                    .at[id, columnName]
                    .replace(
                        oldContent[isaFile[0:1].columns[x]].values[0],
                        sanitizeInput(newContent[x]),
                    )
                )

            # save the changes to the excel file
            try:
                with pd.ExcelWriter(
                    pathName, engine="openpyxl", mode="a", if_sheet_exists="replace"
                ) as writer:
                    isaFile.to_excel(
                        writer, sheet_name=sheetName, merge_cells=False, index=False
                    )
            except:
                raise HTTPException(
                    status_code=500,
                    detail="Error writing the Excel File. Please check your excel file and try to repair it if corrupted!",
                )
        else:
            return "Nothing changed, row not found!"
    # return the name of the row back
    return isaFile.iat[id, 0]


# help function to figure out what isa file we are editing (for the sheet name)
def getIsaType(path: str):
    # split the path into an array with "/" as the separator
    pathSplit = path.split("/")

    # take the top entry of the array
    fileName = pathSplit.pop().lower()

    # check if the file is starting with "isa" and ending with "xlsx"
    if fileName[:3] == "isa" and fileName[-4:] == "xlsx":
        # return the type of the isa
        return fileName.split(".")[1]
    else:
        return ""


# returns a list of all the non metadata sheets and their names
def getSwateSheets(path: str, type: str):
    excelFile = pd.ExcelFile(path)
    sheets = []
    names = []
    match type:
        case "study":
            sheetNames = excelFile.sheet_names
            # if the sheetName is not "Study" or "isa_study", then its a swate sheet
            sheets = [
                loads(
                    pd.read_excel(path, sheet_name=x, engine="openpyxl").to_json(
                        orient="split"
                    )
                )
                for x in sheetNames
                if x != "Study" and x != "isa_study"
            ]

            names = [x for x in sheetNames if x != "Study" and x != "isa_study"]

        case "assay":
            sheetNames = excelFile.sheet_names

            sheets = [
                loads(
                    pd.read_excel(path, sheet_name=x, engine="openpyxl").to_json(
                        orient="split"
                    )
                )
                for x in sheetNames
                if x != "Assay" and x != "isa_assay"
            ]

            names = [x for x in sheetNames if x != "Assay" and x != "isa_assay"]

        case "datamap":
            sheetNames = excelFile.sheet_names

            sheets = [
                loads(
                    pd.read_excel(path, sheet_name=x, engine="openpyxl").to_json(
                        orient="split"
                    )
                )
                for x in sheetNames
            ]

            names = [x for x in sheetNames]

    return sheets, names


# fill a new table column wise with the given data and safe it to the excel file
def createSheet(sheetContent: sheetContent, target: str):
    head = []
    content = []

    tableHead = sheetContent.tableHead
    tableData = sanitizeInput(sheetContent.tableContent)
    path = sheetContent.path
    name = sheetContent.name
    id = sheetContent.id

    # loop column by column
    for i, entry in enumerate(tableHead):
        columnData = [cell for cell in tableData[i]]
        try:
            # if its a custom column, mark it with a [C] at the end
            if entry["Custom"]:
                head.append(sanitizeInput(str(entry["Type"])) + "[C]")
            else:
                head.append(sanitizeInput(str(entry["Type"])))
        except:
            head.append(sanitizeInput(str(entry["Type"])))

        # add the data cells for the column to content
        content.append(columnData)

    # create a dataframe using the first column header and data as basis
    df = pd.DataFrame({head[0]: content[0]})

    # remove first column and data to prevent duplication
    head.pop(0)
    content.pop(0)

    # add data column by column to the dataframe
    for i, entry in enumerate(head):
        try:
            df.insert(i + 1, entry, content[i], allow_duplicates=True)
        except:
            pass

    pathName = f"{os.environ.get('BACKEND_SAVE')}{target}-{id}/{path}"

    importIsa = Xlsx.from_xlsx_file(pathName)
    importIsa.RemoveWorksheet(name)

    try:
        # save data to file
        with pd.ExcelWriter(
            pathName, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            df.to_excel(writer, sheet_name=name, index=False)
    except:
        raise HTTPException(
            status_code=500,
            detail="Error writing the Excel File. Please check your excel file and try to repair it in case of corruption!",
        )
    wb = openpyxl.load_workbook(filename=pathName)

    if getIsaType(path) == "datamap":
        # creates a new table inside of the excel sheet
        tab = openpyxl.worksheet.table.Table(
            displayName="datamapTable" + name.replace(" ", "_"),
            ref=f"A1:{openpyxl.utils.get_column_letter(df.shape[1])}{len(df)+1}",
        )
        # alternative table in case of naming clashes
        tab2 = openpyxl.worksheet.table.Table(
            displayName="datamapTable" + name.replace(" ", "_") + "123",
            ref=f"A1:{openpyxl.utils.get_column_letter(df.shape[1])}{len(df)+1}",
        )
    else:
        # creates a new table inside of the excel sheet
        tab = openpyxl.worksheet.table.Table(
            displayName="annotationTable" + name.replace(" ", "_"),
            ref=f"A1:{openpyxl.utils.get_column_letter(df.shape[1])}{len(df)+1}",
        )

        tab2 = openpyxl.worksheet.table.Table(
            displayName="annotationTable" + name.replace(" ", "_") + "123",
            ref=f"A1:{openpyxl.utils.get_column_letter(df.shape[1])}{len(df)+1}",
        )

    # styles an excel table sometimes similar to swate
    style = openpyxl.worksheet.table.TableStyleInfo(
        name="TableStyleMedium11",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    tab.tableStyleInfo = style
    tab2.tableStyleInfo = style
    try:
        wb[name].add_table(tab)
    except:
        try:
            wb[name].add_table(tab2)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"ERROR: {e}")
    wb.save(pathName)


# when you sync an assay to a study, either overwrite existing data or append new data
def appendAssay(pathToAssay: str, pathToStudy: str, assayName: str):
    # parse the correct sheet (first 8 rows for assay files)
    try:
        assay = pd.read_excel(
            pathToAssay, sheet_name="isa_assay", engine="openpyxl"
        ).head(8)
    except:
        try:
            assay = pd.read_excel(
                pathToAssay, sheet_name="Assay", engine="openpyxl"
            ).head(8)
        except:
            assay = pd.read_excel(pathToAssay, 0, engine="openpyxl").head(8)

    try:
        study = pd.read_excel(pathToStudy, sheet_name="isa_study", engine="openpyxl")
        sheetName = "isa_study"
    except:
        try:
            study = pd.read_excel(pathToStudy, sheet_name="Study", engine="openpyxl")
            sheetName = "Study"
        except:
            study = pd.read_excel(pathToStudy, 0, engine="openpyxl")
            sheetName = ""

    # fill nan with empty strings
    assay = assay.fillna("")
    study = study.fillna("")

    # find row containing the assay data in the study file
    try:
        assayIndex = (
            study.index[study["STUDY METADATA"] == "STUDY ASSAYS"].to_list()[0] + 1
        )
    except:
        try:
            assayIndex = study.index[study["STUDY"] == "STUDY ASSAYS"].to_list()[0] + 1
        except:
            raise HTTPException(
                status_code=400,
                detail="Study has no STUDY ASSAYS field",
            )

    # the number of columns
    columnLength = len(study.columns.to_list())

    # if there is just one column, add a second one to make space for the assay data
    if columnLength == 1:
        study["Unnamed: 1"] = ""

    # if the assay has no data and therefore no second column, add one
    assayColumns = len(assay.columns.to_list())

    try:
        if assayColumns == 1:
            assay.insert(assayColumns, "Unnamed: 1", "")
    except:
        print("No new column could be added to the assay!")

    # index of the free column
    freeColumn = 0

    # set to true if the same assay is already in the study
    overwrite = assayName in study.iloc[[assayIndex + 7]].to_string()

    # check for a free column
    for x in range(1, columnLength):
        # if assay already exits, get the column index
        if overwrite:
            if assayName in study.iat[assayIndex + 7, x]:
                freeColumn = x
                break

        # check the ASSAY FILE NAME Field for free space to insert the new assay
        elif study.iat[assayIndex + 7, x] == "":
            freeColumn = x
            break

    # if there is no free column add a new empty column
    if freeColumn == 0:
        freeColumn = len(study.columns)
        study["Unnamed: " + str(freeColumn)] = ""

    # fill the column with the assay data
    for x in range(len(assay)):
        study.iat[assayIndex + x, freeColumn] = assay.iat[x, 1]

    try:
        # save the changes to the excel file
        with pd.ExcelWriter(
            pathToStudy, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            study.to_excel(writer, sheet_name=sheetName, merge_cells=False, index=False)
    except:
        raise HTTPException(
            status_code=500,
            detail="Error writing the Excel File. Please check your excel file and try to repair it if corrupted!",
        )
    return study.to_json()


# append study to investigation file
def appendStudy(pathToStudy: str, pathToInvest: str, studyName: str):
    # parse the correct sheet
    try:
        invest = pd.read_excel(
            pathToInvest, sheet_name="isa_investigation", engine="openpyxl"
        )
        sheetName = "isa_investigation"
    except:
        invest = pd.read_excel(pathToInvest, 0, engine="openpyxl")
        sheetName = ""

    try:
        study = pd.read_excel(pathToStudy, sheet_name="isa_study", engine="openpyxl")
    except:
        try:
            study = pd.read_excel(pathToStudy, sheet_name="Study", engine="openpyxl")
        except:
            study = pd.read_excel(pathToStudy, 0, engine="openpyxl")

    # fill nan with empty strings
    invest = invest.fillna("")
    study = study.fillna("")

    # get a list of all rows named "Study Identifier"
    try:
        studyIndex = invest.index[
            invest["ONTOLOGY SOURCE REFERENCE"] == "Study Identifier"
        ].to_list()
    except:
        raise HTTPException(
            status_code=400, detail="Investigation file has no Study Identifier"
        )

    # index of the row containing the already existing study/the next free available row
    rowIndex = 0

    # find the right row index (first check if study is already there, or take the next free space)
    for x in studyIndex:
        if studyName in invest.iloc[[x]].to_string().replace(" ", ""):
            rowIndex = x
            break
        # check for free space, but prefer already exiting study
        if invest.iloc[x, 1] == "" and rowIndex == 0:
            rowIndex = x

    # if the study contains more columns, then extend the investigation file by the amount of missing columns
    if len(invest.columns) < len(study.columns):
        for x in range(len(study.columns) - len(invest.columns)):
            invest["Unnamed: " + str(len(invest.columns) + x)] = ""

    # if the study is not yet in the investigation file and there is also no free space, append a new empty study and fill it with the data
    if rowIndex == 0:
        # get a sample of an empty study
        emptyStudy = pd.read_excel(
            os.environ.get("BACKEND_SAVE") + "isa_files/isa.study.xlsx",
            engine="openpyxl",
        )

        # rename the first column to match the column name of the investigation file (or else it will be added as a new column)
        emptyStudy.rename(
            columns={"STUDY METADATA": "ONTOLOGY SOURCE REFERENCE"}, inplace=True
        )
        # the index of the row to start is at the bottom of the invest file
        rowIndex = len(invest) + 1

        # add the row "STUDY" to indicate a new study
        invest.loc[len(invest)] = {"ONTOLOGY SOURCE REFERENCE": "STUDY"}

        # add the empty study to the investigation file
        invest = pd.concat([invest, emptyStudy], ignore_index=True)

    # finally fill the data in the correct rows
    for x in range(len(study)):
        for y in range(len(study.columns)):
            try:
                invest.iat[rowIndex + x, y] = study.iat[x, y]
            except Exception as e:
                invest.loc[len(invest)] = {"ONTOLOGY SOURCE REFERENCE": "OTHER"}
                invest.iat[rowIndex + x, y] = study.iat[x, y]

    try:
        # save the changes to the excel file
        with pd.ExcelWriter(
            pathToInvest, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            invest.to_excel(
                writer, sheet_name=sheetName, merge_cells=False, index=False
            )
    except:
        raise HTTPException(
            status_code=500,
            detail="Error writing the Excel File. Please check your excel file and try to repair it if corrupted!",
        )

    return invest.to_json()


# reads out the given file and sends the content as json back
def readExcelFile(file: bytes):
    # initiate isaFile structure
    excelFile: pd.DataFrame

    # read the file
    try:
        excelFile = pd.read_excel(file, engine="openpyxl")
    except:
        excelFile = pd.read_excel(file, 0, engine="openpyxl")

    # parse the dataframe into json and return it
    parsed = loads(excelFile.to_json(orient="split"))

    return parsed
